from datetime import datetime
import openpyxl

from Gericht import Gericht
from Tag import Tag
from People import People
from Zutat import Zutat

class Excel:
    
    def __init__(self):
        self._file = None
        self._page = None

    def read_file(self, filePath):
        self._file = openpyxl.load_workbook(filePath, data_only=True)
        return self._file

    def read_page(self, worksheet):
        self._page = self._file[worksheet]
        return self._page
        
    def read_people(self):
        _participant = self._page['A2'].value
        _tent = self._page['A3'].value
        _free_man = self._page['A4'].value
        _admin = self._page['A5'].value
        _cook = self._page['A6'].value
        _sum = self._page['A7'].value
        _tent_place = self._page['A8'].value
        _team = self._page['A9'].value

        return People(
            participant=_participant,
            tent=_tent,
            free_man=_free_man,
            admin=_admin,
            cook=_cook,
            sum=_sum,
            tent_place=_tent_place,
            team=_team,
        )
    
    def read_tag(self):

        _people = self.read_people()

        _datum_raw = self._page['C12'].value
        _datum = _datum_raw.strftime("%d.%m.%Y") if isinstance(_datum_raw, datetime) else str(_datum_raw)
        _gericht:Gericht = self.read_gericht()

        _tag = Tag(
            people=_people,
            datum=_datum,
            gericht=_gericht,
            wochentag=self._page.title
            )

        return(_tag)
    
    def read_gericht(self):
        gerichte = []
        _zutaten = []
        mahlzeit = None
        uhrzeit = None
        gerichtname = None

        for row in self._page.iter_rows(min_row=14, max_row=self._page.max_row, min_col=1, max_col=8):
            if row[0].value and row[2].value and row[5].value:
                if mahlzeit and uhrzeit and gerichtname:
                    gericht:Gericht = Gericht(mahlzeit, gerichtname, uhrzeit, _zutaten)
                    gerichte.append(gericht)
                    _zutaten = []

                mahlzeit = row[0].value.split(":")[0]
                uhrzeit = row[2].value
                gerichtname = row[5].value
            
            else:
                zutat = self.read_zutat(row)
                if zutat.artikelname and zutat.menge and zutat.einheit and zutat.lieferant:
                    _zutaten.append(zutat)

        gerichte.append(Gericht(mahlzeit, gerichtname, uhrzeit, _zutaten))

        return gerichte

    def read_zutat(self, row):
        _lieferant = row[1].value
        _kategorie = row[2].value
        _menge = row[3].value
        _einheit = row[4].value
        _artikelname = row[5].value
        _faktor = row[6].value
        _sonstiges = row[7].value

        _zutat:Zutat = Zutat(
            _lieferant,
            _kategorie,
            _menge,
            _einheit,
            _artikelname,
            _faktor,
            _sonstiges)

        return _zutat
    